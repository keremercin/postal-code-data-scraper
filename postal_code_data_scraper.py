from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl

def scrape_postal_code_data(input_excel_path, output_excel_path, base_url):
    # Open the input Excel file
    workbook = openpyxl.load_workbook(input_excel_path)
    worksheet = workbook.active

    # Create a new Excel file and add column headers
    new_workbook = openpyxl.Workbook()
    new_worksheet = new_workbook.active
    new_worksheet.append(["PostalCode", "Value1", "Value2", "Value3"])

    # Initialize the web driver
    driver = webdriver.Chrome()

    # Read postal codes from the input Excel file
    postal_codes = [str(row[0]) for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True)]

    # Process each postal code
    for postal_code in postal_codes[:3]:  # Limiting to the first 3 postal codes
        driver.get(base_url)

        # Scroll down the page to ensure elements are loaded
        for _ in range(3):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)

        # Wait until the search box is available
        wait = WebDriverWait(driver, 10)
        search_box = wait.until(EC.presence_of_element_located((By.ID, "geo-search")))

        # Enter the postal code
        search_box.clear()
        search_box.send_keys(postal_code)

        # Wait until the search button is clickable and click it
        button = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "_chrr2022_geo-search__search-dropdown-link")))
        button.click()

        # Wait until the data row is visible
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "tr[data-component-name='Adult Obesity'] td span")))

        # Extract data from the page
        cells = driver.find_elements(By.CSS_SELECTOR, "tr[data-component-name='Adult Obesity'] td")[1:]
        value1 = cells[0].text if cells else "N/A"
        value2 = cells[1].text if len(cells) > 1 else "N/A"
        value3 = cells[2].text if len(cells) > 2 else "N/A"

        # Append data to the new Excel file
        new_worksheet.append([postal_code, value1, value2, value3])

        # Wait a short time before the next iteration
        time.sleep(2)

    # Save the new Excel file
    new_workbook.save(output_excel_path)

    # Close the web driver
    driver.quit()

if __name__ == "__main__":
    input_excel_path = "posta_kodlari.xlsx"  # Change this to the path of your input Excel file
    output_excel_path = "veriler.xlsx"  # Change this to the path where you want to save the output Excel file
    base_url = "https://www.countyhealthrankings.org/explore-health-rankings/county-health-rankings-model/health-factors/health-behaviors/diet-and-exercise/adult-obesity?year=2022"

    scrape_postal_code_data(input_excel_path, output_excel_path, base_url)
